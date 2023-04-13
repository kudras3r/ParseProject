
"""

University Parser - Парсер ВУЗов (Telegram Bot) - https://t.me/UnivParseBot 
Creator - kudras3r (https://t.me/kudras3r_dev)
Licence - COPYRIGHT ECHO'S DEVELOPMENT © 2023 
All rigts reserved

"""

# Imports
import requests
import openpyxl
import json
import sys

from bs4 import BeautifulSoup
from colorama import Style, Fore, Back
from config import *
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from time import sleep


# User program accept
def select_program() -> str:
    '''
        This is a function to work in the console.
        Accepts a city from the user.
    '''
    global user_program
    
    print('Please select the study program...')
    for program_name in PROGRAMS_ENG.keys():
        print(Fore.CYAN + program_name + Style.RESET_ALL)
    
    user_program = input('>>> ')
    user_program = user_program.lower().capitalize()
    
    if user_program in PROGRAMS_ENG.keys():
        print(PROGRAMS_ENG[user_program])
        return PROGRAMS_ENG[user_program]
    else:
        print(Fore.RED + 'Invalid program name!' + Style.RESET_ALL)
        exit()


# user city 
def select_city() -> str:    
    '''
        This is a function to work in the console.
        Accepts a city from the user.
    '''
    global user_city
    
    # First message
    print(Fore.RED + "You have launched the university parser.\nPlease select the city by which universities will be searched.\nThe result will be send you in 'json' or 'xlsx'" + Style.RESET_ALL)
    print('-'*120)
    print('Choose city. Enter the short name.')
    
    for name, short_name in CITIES_RUS.items():
        print(Fore.WHITE + f'for {name} enter: {Fore.CYAN + short_name + Style.RESET_ALL}')
    
    user_city = input('>>> ')
        
    if user_city.upper() in CITIES_ENG.values():
        return user_city.lower()
    else:
        print(Fore.RED + 'Invalid city code!' + Style.RESET_ALL)
        exit()


# Creating xlsx file
def xlsx_data(data: dict, city=user_city, program=user_program) -> str:

    row1 = 2
    
    ''' Creating xlsx book '''
    book = openpyxl.Workbook()
    sheet = book.active
    
    ''' Fixes xlsx '''
    sheet['A1'] = 'Направление'
    sheet['B1'] = 'Университеты'
    sheet['C1'] = 'Баллы'
    
    sheet['A1'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='0070C0', size=14)
    sheet['B1'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='0070C0', size=14)
    sheet['C1'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='0070C0', size=12)
    
    sheet.column_dimensions['A'].width = 70
    sheet.column_dimensions['B'].width = 70

    for programm in data.keys():
        univ_info = ''
        sheet.cell(row=row1, column= 1).value = programm
        for univ_name, link in data[programm].items():
            if univ_name != 'Баллы на бюджет: ':
                univ_info += f' \n{univ_name}: {link}'
            else:
                sheet[f'B{row1}'].alignment = Alignment(wrapText=True)
                sheet.cell(row=row1, column=3).value = link

        sheet[f'A{row1}'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='9a76f5', size=13)
        sheet[f'B{row1}'].font = Font(name='Arials', charset=204, family=2.0, b=True, color='4f1818', size=12)
        
        sheet.cell(row=row1, column=2).value = univ_info.replace(',', '')

        row1 += 1
  
            
    ''' Saving book in ./results '''
    book.save(f"{city}_result_{program[1:-1]}.xlsx")
    book.close()

    return f"{city}_result_{program[1:-1]}.xlsx"  #file name


# data collection function(url) -> data
def get_data(URL) -> dict:
    '''
        The main function for the parser. collects all information.
    '''
    result_dict = dict()

    # make request to 1st page - (get for taking cookie an posts for sending it)
    s = requests.Session()    # Session
    sleep(0.2)      # for security
    s.get(URL, headers=headers)
    print(f'{headers} - us_agent')
    time_cookies = requests.utils.dict_from_cookiejar(s.cookies)
    
    response = s.post(URL, headers=headers, cookies=time_cookies)
    
    # check status code
    if response.status_code == 200:

        print(f'status code: {response.status_code}')
        print('wait...')
        
        # Create a soup for parse
        soup = BeautifulSoup(response.text, 'lxml')
        
        #    Сhecking the number of pages on the site.
        #    If there is only one page, it goes to else
        
        if 'invite fetcher' in response.text:
            
            navigation_count = int(soup.find('div', class_='invite fetcher').find_all('a', class_='paginator')[-1].text)

            for page in range(1, navigation_count + 1): 
        
                print(f'{page}/{navigation_count}')
            
                url = f'{URL}?page_num={page}'
                response = s.post(url, headers=headers, cookies=time_cookies)
            
                soup = BeautifulSoup(response.text, 'lxml')
        
                all_programs_on_page = soup.find('ul', class_='list-unstyled list-wrap').find_all('div', class_='list__info')
                
                page_iterate(all_programs_on_page, result_dict, time_cookies, s)
        
        # If there are more than one page, iterate over each and collect information
        else:
            url = URL
            response = s.post(url, headers=headers, cookies=time_cookies)
            soup = BeautifulSoup(response.text, 'lxml')
            all_programs_on_page = soup.find('ul', class_='list-unstyled list-wrap').find_all('div', class_='list__info')
            page_iterate(all_programs_on_page, result_dict, time_cookies, s)

        print('\n')
        print(Fore.GREEN + 'Complete!' + Style.RESET_ALL)
        print('-'*120)

        return result_dict    # main data!
    
    else:
        print('Error')
        sys.exit

# pages iterate
def page_iterate(all_programs_on_page, result_dict, time_cookies, s):

    ''' Iterates through the pages from the list all_programs_on_page, 
        places the results in result_dict.
        To work, you need to set cookies and session (s)'''

    for program in all_programs_on_page:

        program_url = program.find('h2', class_='list__h').find('a')['href']
        program_name = program.find('h2', class_='list__h').find('a').text
                
        if program.find('span', class_='list__score-sm') != None:
                    
            min_points = program.find('span', class_='list__score-sm').find('b').text
                    
        else:
            min_points = 'No info'

        if 'vuz' in program_url:
                
            univ_name = program.find('p', class_='list__pre').find('a').text
            univ_link = program.find('h2', class_='list__h').find('a')['href']
                
            result_dict[program_name] = {univ_name: univ_link}
            result_dict[program_name]['Баллы на бюджет: '] = min_points
                    
        else:
            univ_names = []
            univ_links = []

            response = s.post(program_url, headers=headers, cookies=time_cookies)
            soup = BeautifulSoup(response.text, 'lxml')

            univ_info = soup.find('div', class_='detail-box__item detail-box__item_upcase').find_all('a')

            for info in univ_info:
                        univ_names.append(info.text.replace(u"\u00A0", ""))
                        univ_links.append(info['href'])

            result_dict[program_name] = dict(zip(univ_names, univ_links))
            result_dict[program_name]['Баллы на бюджет: '] = min_points


# Function for testing (not working at main code)
def print_result(result: dict):

    for program, univ_list in result.items():
        
        print(Fore.BLUE + '-'*((120-len(program))//2) + Fore.BLUE + program + '-'*((120-len(program))//2))

        for name, link in univ_list.items():

            print(f'{Fore.WHITE + name}: {Fore.CYAN + link}')


# Main function
def json_creating():
    
    city = select_city()
    prog = select_program()
    url = 'https://' + city + '.postupi.online/programmy-obucheniya/bakalavr/razdel' + prog

    data = get_data(url)
    print(data)
    
    ''' Creating json with data'''
    file = open(f'results/{city}_result.json', 'w', encoding='utf-8')
    json.dump(data, file, indent=4, ensure_ascii=False)
    file.close()    

    print(Fore.GREEN + f'Succesfuly! Check {city}_result.json' + Style.RESET_ALL)

# TESTS for console
if __name__ == '__main__':
   

   #data = get_data(select_city())
   #xlsx_data(data, user_city)
   #json_creating()
   prog = select_program()
   city = select_city()
   url = 'https://' + city + '.postupi.online/programmy-obucheniya/bakalavr/razdel' + prog
   #print(url)
   data = get_data(url)
   xlsx_data(data, city, prog)
   

    
